"""Tests du router crypto : comptes d'exchange et synchronisation des trades."""
import datetime
from unittest.mock import patch

import pytest

import encryption
from models import ExchangeAccount, Position, Purchase, Sale


class FakeClient:
    """Client d'exchange factice — remplace Kraken/Bitvavo dans les tests."""

    trades: list = []

    def __init__(self, *_args, **_kwargs):
        pass

    def test_connection(self):
        return "2 actif(s) détecté(s)"

    balances: dict = {}

    def get_balances(self):
        return dict(self.balances)

    def get_trades(self, extra_assets=None):
        return list(self.trades)


def _trade(ref, asset, side, qty, price, day, fee=0.0):
    return {
        "ref": ref, "asset": asset, "quote": "EUR", "side": side,
        "quantity": qty, "unit_price": price, "fee": fee,
        "date": datetime.date(2026, 1, day),
    }


@pytest.fixture
def fake_exchange():
    FakeClient.trades = []
    FakeClient.balances = {}
    with patch("routers.crypto.get_client", lambda *a, **k: FakeClient()):
        yield FakeClient


def _create_account(client, exchange="kraken", label="Kraken test"):
    return client.post("/api/crypto/accounts", json={
        "exchange": exchange, "label": label,
        "api_key": "key", "api_secret": "c2VjcmV0",
    })


# ── Comptes ───────────────────────────────────────────────

def test_create_account_encrypts_keys(auth_admin, db, fake_exchange):
    client, _ = auth_admin
    r = _create_account(client)
    assert r.status_code == 201, r.text

    body = r.json()
    assert body["exchange"] == "kraken"
    # Les clés ne doivent jamais transiter dans la réponse
    assert "api_key" not in body and "api_secret" not in body

    acc = db.query(ExchangeAccount).filter(ExchangeAccount.id == body["id"]).first()
    assert acc.encrypted_api_key != "key"
    assert encryption.decrypt_str(acc.encrypted_api_key) == "key"
    assert encryption.decrypt_str(acc.encrypted_api_secret) == "c2VjcmV0"


def test_create_account_rejects_unknown_exchange(auth_admin, fake_exchange):
    client, _ = auth_admin
    r = client.post("/api/crypto/accounts", json={
        "exchange": "binance", "label": "X", "api_key": "k", "api_secret": "s",
    })
    assert r.status_code == 400
    assert "non supporté" in r.json()["detail"]


def test_regular_user_cannot_create_account(auth_user, fake_exchange):
    r = _create_account(auth_user)
    assert r.status_code in (401, 403)


# ── Synchronisation ───────────────────────────────────────

def test_sync_imports_buys_and_sells(auth_admin, db, fake_exchange):
    client, _ = auth_admin
    acc_id = _create_account(client).json()["id"]

    fake_exchange.trades = [
        _trade("t1", "BTC", "buy",  0.5, 40000.0, 10, fee=5.0),
        _trade("t2", "BTC", "buy",  0.5, 50000.0, 12),
        _trade("t3", "BTC", "sell", 0.4, 60000.0, 20),
    ]
    r = client.post(f"/api/crypto/accounts/{acc_id}/sync", json={})
    assert r.status_code == 200, r.text
    assert r.json()["created_purchases"] == 2
    assert r.json()["created_sales"] == 1

    pos = db.query(Position).filter(Position.asset_type == "crypto").one()
    assert pos.ticker == "BTC-EUR"          # format yfinance
    assert pos.display_name == "BTC · Kraken test"

    sale = db.query(Sale).filter(Sale.position_id == pos.id).one()
    # PRU = (0.5*40000 + 5 + 0.5*50000) / 1.0 = 45005
    assert sale.realized_pnl == pytest.approx((60000 - 45005) * 0.4)


def test_sync_is_idempotent(auth_admin, db, fake_exchange):
    """Re-synchroniser ne doit pas dupliquer les transactions déjà importées."""
    client, _ = auth_admin
    acc_id = _create_account(client).json()["id"]
    fake_exchange.trades = [_trade("t1", "ETH", "buy", 2.0, 2000.0, 5)]

    first = client.post(f"/api/crypto/accounts/{acc_id}/sync", json={}).json()
    assert first["created_purchases"] == 1 and first["skipped"] == 0

    second = client.post(f"/api/crypto/accounts/{acc_id}/sync", json={}).json()
    assert second["created_purchases"] == 0 and second["skipped"] == 1

    assert db.query(Purchase).count() == 1


def test_sync_archives_fully_sold_position(auth_admin, db, fake_exchange):
    client, _ = auth_admin
    acc_id = _create_account(client).json()["id"]
    fake_exchange.trades = [
        _trade("b1", "SOL", "buy",  10.0, 100.0, 3),
        _trade("s1", "SOL", "sell", 10.0, 150.0, 8),
    ]
    client.post(f"/api/crypto/accounts/{acc_id}/sync", json={})

    pos = db.query(Position).filter(Position.asset_type == "crypto").one()
    assert pos.is_active is False


def test_sync_skips_non_eur_pairs(auth_admin, db, fake_exchange):
    client, _ = auth_admin
    acc_id = _create_account(client).json()["id"]
    usd = _trade("u1", "BTC", "buy", 1.0, 40000.0, 4)
    usd["quote"] = "USD"
    fake_exchange.trades = [usd]

    r = client.post(f"/api/crypto/accounts/{acc_id}/sync", json={}).json()
    assert r["created_purchases"] == 0
    assert r["unsupported_pairs"] == ["BTC/USD"]


def test_separate_positions_per_account(auth_admin, db, fake_exchange):
    """Le même actif sur deux exchanges garde deux PRU distincts."""
    client, _ = auth_admin
    kraken = _create_account(client, "kraken", "Kraken test").json()["id"]
    bitvavo = _create_account(client, "bitvavo", "Bitvavo test").json()["id"]

    fake_exchange.trades = [_trade("k1", "BTC", "buy", 1.0, 30000.0, 2)]
    client.post(f"/api/crypto/accounts/{kraken}/sync", json={})
    fake_exchange.trades = [_trade("b1", "BTC", "buy", 1.0, 60000.0, 2)]
    client.post(f"/api/crypto/accounts/{bitvavo}/sync", json={})

    names = {p.display_name for p in db.query(Position)
             .filter(Position.asset_type == "crypto").all()}
    assert names == {"BTC · Kraken test", "BTC · Bitvavo test"}


# ── Vue d'ensemble ────────────────────────────────────────

def test_summary_excludes_crypto_from_portfolio(auth_admin, db, fake_exchange):
    """Le crypto a sa propre page : il ne doit pas polluer /api/portfolio/summary."""
    client, _ = auth_admin
    acc_id = _create_account(client).json()["id"]
    fake_exchange.trades = [_trade("t1", "BTC", "buy", 1.0, 30000.0, 2)]
    client.post(f"/api/crypto/accounts/{acc_id}/sync", json={})

    with patch("market_data.get_price_eur", return_value=(50000.0, 49000.0, 2.0)):
        crypto = client.get("/api/crypto/summary").json()
        portfolio = client.get("/api/portfolio/summary").json()

    assert crypto["stats"]["position_count"] == 1
    assert crypto["stats"]["total_value_eur"] == pytest.approx(50000.0)
    assert all(p["asset_type"] != "crypto" for p in portfolio["positions"])


# ── Liquidités d'exchange ─────────────────────────────────

def test_sync_records_eur_cash_balance(auth_admin, db, fake_exchange):
    """L'argent non investi qui dort sur l'exchange doit remonter aussi."""
    client, _ = auth_admin
    acc_id = _create_account(client, "bitvavo", "Bitvavo - Bot").json()["id"]
    fake_exchange.trades = [_trade("t1", "BTC", "buy", 1.0, 100.0, 2)]
    fake_exchange.balances = {"EUR": 282.92, "BTC": 1.0}

    r = client.post(f"/api/crypto/accounts/{acc_id}/sync", json={}).json()
    assert r["cash_balance_eur"] == pytest.approx(282.92)

    with patch("market_data.get_price_eur", return_value=(217.19, 217.0, 0.1)):
        summary = client.get("/api/crypto/summary").json()
    assert summary["stats"]["total_cash_eur"] == pytest.approx(282.92)
    # Capital total = investi + liquidités (les 500 € du compte Bitvavo)
    assert summary["stats"]["total_account_eur"] == pytest.approx(217.19 + 282.92)
    assert summary["cash"][0]["display_name"] == "Liquidités · Bitvavo - Bot"


def test_holdings_without_trades_are_reconciled(auth_admin, db, fake_exchange):
    """Un avoir arrivé par dépôt (aucun trade) doit quand même apparaître."""
    client, _ = auth_admin
    acc_id = _create_account(client, "kraken", "Kraken").json()["id"]
    fake_exchange.trades = []                     # rien n'a été acheté sur Kraken
    fake_exchange.balances = {"BTC": 0.25, "ETH": 2.0, "EUR": 0.0}

    with patch("market_data.get_price_eur", return_value=(40000.0, 39000.0, 1.0)):
        r = client.post(f"/api/crypto/accounts/{acc_id}/sync", json={}).json()
    assert r["created_purchases"] == 0            # aucun trade importé
    assert set(r["reconciled_assets"]) == {"BTC", "ETH"}

    btc = db.query(Position).filter(Position.display_name == "BTC · Kraken").one()
    assert btc.is_active is True
    assert sum(p.quantity for p in btc.purchases) == pytest.approx(0.25)


def test_reconciliation_only_covers_the_untracked_part(auth_admin, db, fake_exchange):
    """Si des trades expliquent une partie du solde, seul l'écart est ajouté."""
    client, _ = auth_admin
    acc_id = _create_account(client, "kraken", "Kraken").json()["id"]
    fake_exchange.trades = [_trade("t1", "BTC", "buy", 0.10, 30000.0, 3)]
    fake_exchange.balances = {"BTC": 0.25}        # 0,10 tradé + 0,15 déposé

    with patch("market_data.get_price_eur", return_value=(40000.0, 39000.0, 1.0)):
        client.post(f"/api/crypto/accounts/{acc_id}/sync", json={})

    btc = db.query(Position).filter(Position.display_name == "BTC · Kraken").one()
    assert sum(p.quantity for p in btc.purchases) == pytest.approx(0.25)
    extra = [p for p in btc.purchases if p.note.endswith(":BALANCE]")]
    assert len(extra) == 1 and extra[0].quantity == pytest.approx(0.15)
    # le PRU issu du trade réel n'est pas écrasé
    assert any(p.unit_price == 30000.0 for p in btc.purchases)


def test_reconciliation_is_idempotent(auth_admin, db, fake_exchange):
    """Resynchroniser ne doit pas empiler des lignes de réconciliation."""
    client, _ = auth_admin
    acc_id = _create_account(client, "kraken", "Kraken").json()["id"]
    fake_exchange.balances = {"BTC": 0.25}
    with patch("market_data.get_price_eur", return_value=(40000.0, 39000.0, 1.0)):
        client.post(f"/api/crypto/accounts/{acc_id}/sync", json={})
        client.post(f"/api/crypto/accounts/{acc_id}/sync", json={})

    btc = db.query(Position).filter(Position.display_name == "BTC · Kraken").one()
    assert len(btc.purchases) == 1
    assert btc.purchases[0].quantity == pytest.approx(0.25)


def test_same_asset_on_two_exchanges_stays_on_separate_lines(auth_admin, db, fake_exchange):
    """Le BTC Kraken et le BTC du bot Bitvavo doivent rester deux lignes."""
    client, _ = auth_admin
    kraken = _create_account(client, "kraken", "Kraken").json()["id"]
    bitvavo = _create_account(client, "bitvavo", "Bitvavo - Bot").json()["id"]

    with patch("market_data.get_price_eur", return_value=(40000.0, 39000.0, 1.0)):
        fake_exchange.balances = {"BTC": 0.25}
        client.post(f"/api/crypto/accounts/{kraken}/sync", json={})
        fake_exchange.balances = {"BTC": 0.0015}
        client.post(f"/api/crypto/accounts/{bitvavo}/sync", json={})

        summary = client.get("/api/crypto/summary").json()

    names = {p["display_name"]: p["total_quantity"] for p in summary["positions"]}
    assert names["BTC · Kraken"] == pytest.approx(0.25)
    assert names["BTC · Bitvavo - Bot"] == pytest.approx(0.0015)


def test_no_eur_on_account_is_zero_not_an_error(auth_admin, db, fake_exchange):
    """Un compte sans euros vaut 0 € — ce n'est pas un problème de clé API."""
    client, _ = auth_admin
    acc_id = _create_account(client, "kraken", "Kraken").json()["id"]
    fake_exchange.balances = {"BTC": 1.0, "ETH": 2.0}   # aucun EUR détenu

    r = client.post(f"/api/crypto/accounts/{acc_id}/sync", json={}).json()
    assert r["cash_error"] is None
    assert r["cash_balance_eur"] == 0.0
    # et aucune ligne « Liquidités » à 0 € n'encombre le tableau
    assert db.query(Position).filter(Position.ticker == "KRAKEN:EUR").count() == 0


def test_unreadable_balances_are_reported(auth_admin, db, fake_exchange):
    """En revanche, un appel qui échoue doit être signalé, pas masqué en 0 €."""
    client, _ = auth_admin
    acc_id = _create_account(client, "kraken", "Kraken").json()["id"]

    def boom(self):
        raise RuntimeError("EAPI:Invalid key")

    with patch.object(FakeClient, "get_balances", boom):
        r = client.post(f"/api/crypto/accounts/{acc_id}/sync", json={}).json()
    assert r["cash_balance_eur"] is None
    assert "Invalid key" in r["cash_error"]
    accounts = client.get("/api/crypto/accounts").json()
    assert "soldes illisibles" in accounts[0]["last_sync_status"]


def test_cash_balance_is_updated_not_duplicated(auth_admin, db, fake_exchange):
    client, _ = auth_admin
    acc_id = _create_account(client, "bitvavo", "Bitvavo - Bot").json()["id"]
    fake_exchange.balances = {"EUR": 100.0}
    client.post(f"/api/crypto/accounts/{acc_id}/sync", json={})
    fake_exchange.balances = {"EUR": 250.0}
    client.post(f"/api/crypto/accounts/{acc_id}/sync", json={})

    cash = db.query(Position).filter(Position.ticker == "BITVAVO:EUR").all()
    assert len(cash) == 1                      # mise à jour, pas de doublon
    assert cash[0].manual_price == pytest.approx(250.0)
    assert db.query(Purchase).filter(Purchase.position_id == cash[0].id).count() == 1


def test_scheduler_refreshes_cash_without_reimporting_trades(auth_admin, db, fake_exchange):
    """Le scheduler doit rafraîchir les soldes, sans rejouer l'import des trades."""
    from routers.crypto import refresh_cash_balances
    client, _ = auth_admin
    acc_id = _create_account(client, "bitvavo", "Bitvavo - Bot").json()["id"]
    fake_exchange.trades = [_trade("t1", "BTC", "buy", 1.0, 100.0, 2)]
    fake_exchange.balances = {"EUR": 100.0}
    client.post(f"/api/crypto/accounts/{acc_id}/sync", json={})

    calls = {"trades": 0}
    original = FakeClient.get_trades
    def counting(self, extra_assets=None):
        calls["trades"] += 1
        return original(self, extra_assets)

    fake_exchange.balances = {"EUR": 250.0}
    with patch.object(FakeClient, "get_trades", counting):
        assert refresh_cash_balances(db) == 1

    cash = db.query(Position).filter(Position.ticker == "BITVAVO:EUR").one()
    assert cash.manual_price == pytest.approx(250.0)
    assert calls["trades"] == 0          # aucun import de trades déclenché


def test_exchange_cash_excluded_from_securities_portfolio(auth_admin, db, fake_exchange):
    """Les liquidités Bitvavo ne doivent pas polluer le portefeuille titres."""
    client, _ = auth_admin
    acc_id = _create_account(client, "bitvavo", "Bitvavo - Bot").json()["id"]
    fake_exchange.balances = {"EUR": 282.92}
    client.post(f"/api/crypto/accounts/{acc_id}/sync", json={})

    with patch("market_data.get_price_eur", return_value=(1.0, 1.0, 0.0)):
        portfolio = client.get("/api/portfolio/summary").json()
        dashboard = client.get("/api/dashboard").json()

    assert all(p["ticker"] != "BITVAVO:EUR" for p in portfolio["positions"])
    # Comptabilisé côté crypto, et présent dans le patrimoine total
    assert dashboard["crypto"]["total_value_eur"] == pytest.approx(282.92)
    assert dashboard["grand_total_eur"] >= 282.92


def test_archived_position_stays_archived_after_resync(auth_admin, db, fake_exchange):
    """Une position archivée à la main ne doit pas réapparaître à la resynchro.

    C'est ce que promet la boîte de dialogue d'archivage sur la page Crypto.
    """
    client, _ = auth_admin
    acc_id = _create_account(client).json()["id"]
    fake_exchange.trades = [_trade("t1", "BTC", "buy", 1.0, 30000.0, 2)]
    client.post(f"/api/crypto/accounts/{acc_id}/sync", json={})

    pos = db.query(Position).filter(Position.asset_type == "crypto").one()
    assert client.delete(f"/api/portfolio/positions/{pos.id}").status_code == 200

    # Resynchro : tous les trades sont déjà connus → rien ne doit changer
    r = client.post(f"/api/crypto/accounts/{acc_id}/sync", json={}).json()
    assert r["created_purchases"] == 0 and r["skipped"] == 1
    db.refresh(pos)
    assert pos.is_active is False

    # En revanche, un nouvel achat la réactive
    fake_exchange.trades.append(_trade("t2", "BTC", "buy", 0.5, 31000.0, 9))
    client.post(f"/api/crypto/accounts/{acc_id}/sync", json={})
    db.refresh(pos)
    assert pos.is_active is True


# ── Exports ───────────────────────────────────────────────

def _seed_crypto(client, fake_exchange):
    acc_id = _create_account(client).json()["id"]
    fake_exchange.trades = [_trade("t1", "BTC", "buy", 1.0, 30000.0, 2)]
    client.post(f"/api/crypto/accounts/{acc_id}/sync", json={})


def test_export_data_separates_crypto_totals(auth_admin, db, fake_exchange):
    """Le crypto doit être compté à part, sans disparaître du patrimoine total."""
    from routers.export import _collect_data
    client, _ = auth_admin
    _seed_crypto(client, fake_exchange)

    with patch("market_data.get_price_eur", return_value=(50000.0, 49000.0, 2.0)):
        data = _collect_data(db, db.query(__import__("models").User).first())

    s = data["summary"]
    assert s["total_crypto_eur"] == pytest.approx(50000.0)
    assert s["total_portfolio_eur"] == 0.0        # aucune action/ETF
    assert s["grand_total_eur"] == pytest.approx(50000.0)   # mais bien dans le total

    crypto_rows = [p for p in data["portfolio"] if p["type"] == "crypto"]
    assert len(crypto_rows) == 1
    assert crypto_rows[0]["purchases"]           # l'historique suit


def test_export_html_contains_crypto_tab(auth_admin, fake_exchange):
    client, _ = auth_admin
    _seed_crypto(client, fake_exchange)

    with patch("market_data.get_price_eur", return_value=(50000.0, 49000.0, 2.0)):
        r = client.post("/api/export", json={"passphrase": "TestPassphrase123!"})
    assert r.status_code == 200, r.text
    html = r.text
    assert "tab-crypto" in html and "renderCrypto" in html


def test_export_excel_has_crypto_sheet(auth_admin, fake_exchange):
    import openpyxl
    from io import BytesIO
    client, _ = auth_admin
    _seed_crypto(client, fake_exchange)

    with patch("market_data.get_price_eur", return_value=(50000.0, 49000.0, 2.0)):
        r = client.get("/api/export/patrimoine-excel")
    assert r.status_code == 200, r.text

    wb = openpyxl.load_workbook(BytesIO(r.content))
    assert "Crypto" in wb.sheetnames
    assert wb["Crypto"].max_row == 2                      # en-tête + 1 ligne
    assert wb["Crypto"].cell(2, 1).value == "BTC"
    assert wb["Crypto"].cell(2, 2).value == "Kraken test"

    # La synthèse doit inclure le crypto dans le patrimoine total
    ws = wb["Synthèse"]
    headers = [c.value for c in ws[1]]
    idx = headers.index("Total crypto (€)") + 1
    assert ws.cell(2, idx).value == pytest.approx(50000.0)
    assert ws.cell(2, headers.index("Patrimoine total (€)") + 1).value == pytest.approx(50000.0)


def test_delete_account_keeps_positions(auth_admin, db, fake_exchange):
    client, _ = auth_admin
    acc_id = _create_account(client).json()["id"]
    fake_exchange.trades = [_trade("t1", "BTC", "buy", 1.0, 30000.0, 2)]
    client.post(f"/api/crypto/accounts/{acc_id}/sync", json={})

    assert client.delete(f"/api/crypto/accounts/{acc_id}").status_code == 200
    assert db.query(ExchangeAccount).count() == 0
    assert db.query(Position).filter(Position.asset_type == "crypto").count() == 1
